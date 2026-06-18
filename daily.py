import argparse
import os
import shutil
import sys
import urllib.parse
from datetime import datetime
from functools import partial

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import pandas as pd
from copy import copy
from openpyxl import load_workbook
from sqlalchemy import create_engine

from lac_registry import (
    BSC_PASTE_CLEAR_THROUGH_COL,
    BSC_PIVOT_COLUMNS,
    PSR_ROLLING_DAYS,
    PSR_YEARLY_DAYS,
    RNC_PASTE_CLEAR_THROUGH_COL,
    RNC_PIVOT_COLUMNS,
    exclude_today_rows,
    fetch_daily_psr_bsc,
    fetch_daily_psr_rnc,
    fetch_overall_psr_network,
    fetch_psr_per_lac,
    fetch_psr_per_service,
    sql_date_window_rolling,
    trim_pivot_columns,
)
from workbook_layout import (
    BSC_ROW,
    HIDE_SHEETS,
    LAC_COL,
    LAC_ROW,
    PSR_AVAIL_COL_DATE,
    PSR_AVAIL_COL_PSR,
    PSR_AVAIL_MAX_ROWS,
    PSR_AVAIL_ROW,
    RNC_ROW,
    SERVICE_COL,
    SERVICE_ROW,
    SHEET_DASHBOARD,
    SHEET_DATA,
    SHEET_PSR_AVAIL,
    TECH_COL,
    TECH_ROW,
    TEMPLATE_GRAPHS,
)

DEFAULT_SQL_CONN = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=TIGO;"
    "DATABASE=GSM_KPI;"
    "UID=pmuser;"
    "PWD=Pmuser@123;"
)

OVERALL_PSR_SQL = f"""
SELECT
    CAST([Result Time] AS DATE) AS [Row Labels],
    CAST(100.0 * (
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from A Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Repeated Paging Responses from A Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from Iu Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Repeated Paging Responses from Iu Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from SGs Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Second Paging Responses from SGs Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Third Paging Responses from SGs Interface] AS BIGINT),0))
    ) / NULLIF((
        SUM(ISNULL(TRY_CAST([Number of First Pagings to A Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of First Pagings to Iu Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of First Pagings to SGs Interface] AS BIGINT),0))
    ), 0) AS DECIMAL(10,3)) AS Overall_PSR,
    CAST(100.0 * (
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from A Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Repeated Paging Responses from A Interface] AS BIGINT),0))
    ) / NULLIF(SUM(ISNULL(TRY_CAST([Number of First Pagings to A Interface] AS BIGINT),0)), 0) AS DECIMAL(10,3)) AS PSR_2G,
    CAST(100.0 * (
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from Iu Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Repeated Paging Responses from Iu Interface] AS BIGINT),0))
    ) / NULLIF(SUM(ISNULL(TRY_CAST([Number of First Pagings to Iu Interface] AS BIGINT),0)), 0) AS DECIMAL(10,3)) AS PSR_3G,
    CAST(100.0 * (
        SUM(ISNULL(TRY_CAST([Number of First Paging Responses from SGs Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Second Paging Responses from SGs Interface] AS BIGINT),0)) +
        SUM(ISNULL(TRY_CAST([Number of Third Paging Responses from SGs Interface] AS BIGINT),0))
    ) / NULLIF(SUM(ISNULL(TRY_CAST([Number of First Pagings to SGs Interface] AS BIGINT),0)), 0) AS DECIMAL(10,3)) AS PSR_4G
FROM [CS_KPIS].[dbo].[host03_pmresult_1912976119]
WHERE {sql_date_window_rolling(PSR_ROLLING_DAYS)}
GROUP BY CAST([Result Time] AS DATE)
ORDER BY [Row Labels] ASC;
"""


def _root() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _template() -> str:
    p = os.path.join(_root(), TEMPLATE_GRAPHS)
    if os.path.isfile(p):
        return p
    alt = os.path.join(_root(), "YAS TZ CS Core_Daily Paging Performance Snapshot.xlsx")
    if os.path.isfile(alt):
        return alt
    parent = os.path.join(os.path.dirname(_root()), "Paging", TEMPLATE_GRAPHS)
    if os.path.isfile(parent):
        return parent
    return p


def _output() -> str:
    d = os.path.join(_root(), "Reports")
    os.makedirs(d, exist_ok=True)
    t = datetime.now().strftime("%H%M")
    return os.path.join(d, f"PSR_Graphs_{t}.xlsx")


def _fetch(sql_conn: str, sql: str) -> pd.DataFrame:
    q = urllib.parse.quote_plus(sql_conn)
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={q}", fast_executemany=True)
    with engine.connect() as c:
        return pd.read_sql(sql, c)


def _v(x):
    if pd.isna(x):
        return None
    if isinstance(x, (pd.Timestamp, datetime)):
        return x.date()
    return x


def _set(ws, row, col, val, anchor=None):
    cell = ws.cell(row=row, column=col, value=_v(val))
    if anchor and not cell.has_style:
        s = ws.cell(row=anchor, column=col)
        if s.has_style:
            cell.font = copy(s.font)
            cell.number_format = copy(s.number_format)


def _paste(ws, df, row, col=1, headers=True):
    if df.empty:
        return 0
    n = 0
    start = row
    if headers:
        for j, c in enumerate(df.columns):
            _set(ws, row, col + j, str(c), row)
            n += 1
        start = row + 1
    for i, tup in enumerate(df.itertuples(index=False)):
        for j, val in enumerate(tup):
            _set(ws, start + i, col + j, val, row)
            n += 1
    return n


def _ensure_psr_avail_sheet(wb):
    if SHEET_PSR_AVAIL in wb.sheetnames:
        return wb[SHEET_PSR_AVAIL]
    ws = wb.create_sheet(SHEET_PSR_AVAIL)
    _set(ws, 1, PSR_AVAIL_COL_DATE, "Date")
    _set(ws, 1, PSR_AVAIL_COL_PSR, "PSR (%)")
    _set(ws, 1, 3, "AVAIL (%)")
    return ws


def _paste_psr_avail(wb, df):
    """Yearly PSR in col B only — col C AVAIL stays for you."""
    ws = _ensure_psr_avail_sheet(wb)
    df = exclude_today_rows(df, "Row Labels").sort_values("Row Labels")
    _set(ws, 1, PSR_AVAIL_COL_DATE, "Date")
    _set(ws, 1, PSR_AVAIL_COL_PSR, "PSR (%)")
    _set(ws, 1, 3, "AVAIL (%)")
    n = 0
    for i, row in enumerate(df.itertuples(index=False)):
        if i >= PSR_AVAIL_MAX_ROWS:
            break
        r = PSR_AVAIL_ROW + i
        _set(ws, r, PSR_AVAIL_COL_DATE, row[0], PSR_AVAIL_ROW)
        _set(ws, r, PSR_AVAIL_COL_PSR, row[1], PSR_AVAIL_ROW)
        n += 2
    return n


def run(sql_conn: str, template: str, out: str, yearly_days: int) -> None:
    shutil.copy2(template, out)
    wb = load_workbook(out)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb.worksheets[0]
    ok = 0

    for label, fn, prep, row, col, hdr in (
        ("RNC", fetch_daily_psr_rnc, lambda d: trim_pivot_columns(exclude_today_rows(d, "Row Labels"), RNC_PIVOT_COLUMNS), RNC_ROW, 1, True),
        ("BSC", fetch_daily_psr_bsc, lambda d: trim_pivot_columns(exclude_today_rows(d, "Row Labels"), BSC_PIVOT_COLUMNS), BSC_ROW, 1, True),
        ("LAC", fetch_psr_per_lac, lambda d: d, LAC_ROW, LAC_COL, True),
    ):
        print(f"  {label}...", end=" ", flush=True)
        df = fn(sql_conn)
        if df.empty:
            print("no data")
            continue
        _paste(ws, prep(df), row, col, hdr)
        print(f"{len(df)} rows")
        ok += 1

    print("  Technology PSR...", end=" ", flush=True)
    df = _fetch(sql_conn, OVERALL_PSR_SQL)
    if not df.empty:
        df = exclude_today_rows(df, "Row Labels").rename(
            columns={"Overall_PSR": "Overall PSR %", "PSR_2G": "2G PSR %", "PSR_3G": "3G PSR %", "PSR_4G": "4G PSR %"}
        )
        _paste(ws, df, TECH_ROW, TECH_COL, True)
        print(f"{len(df)} rows")
        ok += 1
    else:
        print("no data")

    print("  Service PSR...", end=" ", flush=True)
    df = fetch_psr_per_service(sql_conn)
    if not df.empty:
        df = exclude_today_rows(df, "Row Labels").rename(
            columns={"PSR_Calls": "Calls PSR %", "PSR_sms": "SMS PSR %", "PSR_ussd": "USSD PSR %", "PSR_psi": "PSI PSR %"}
        )
        _paste(ws, df, SERVICE_ROW, SERVICE_COL, True)
        print(f"{len(df)} rows")
        ok += 1
    else:
        print("no data")

    print("  Yearly PSR (AVAIL manual)...", end=" ", flush=True)
    df = fetch_overall_psr_network(sql_conn, rolling_days=yearly_days)
    if not df.empty:
        _paste_psr_avail(wb, df)
        print(f"{len(df)} rows")
        ok += 1
    else:
        print("no data")

    for name in HIDE_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    if SHEET_DASHBOARD in wb.sheetnames:
        wb.active = wb[SHEET_DASHBOARD]

    wb.save(out)
    mb = os.path.getsize(out) / 1024 / 1024
    print(f"\nDone — {ok} blocks | {mb:.1f} MB | opens fast")
    print(f"Graphs tab: {SHEET_DASHBOARD}")
    print(f"Fill AVAIL (%) manually on sheet: {SHEET_PSR_AVAIL} column C")
    print(out)


def main():
    p = argparse.ArgumentParser(description="PSR graphs — light xlsx, SQL auto-fill.")
    p.add_argument("--template", default="")
    p.add_argument("--out", default="")
    p.add_argument("--yearly-days", type=int, default=PSR_YEARLY_DAYS)
    p.add_argument("--sql-conn", default="")
    args = p.parse_args()

    template = args.template or _template()
    if not os.path.isfile(template):
        sys.exit(f"Missing template: {TEMPLATE_GRAPHS}\nPut it in: {_root()}")

    sql = args.sql_conn or DEFAULT_SQL_CONN or os.environ.get("SQL_CONN", "")
    if not sql:
        sys.exit("No SQL connection string.")

    out = args.out or _output()
    print("PSR graphs export")
    print(f"  Template: {template}")
    print(f"  Output:   {out}")
    run(sql, template, out, args.yearly_days)


if __name__ == "__main__":
    main()
