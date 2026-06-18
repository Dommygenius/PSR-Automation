"""
Extract valuable PSR data into a new consolidated sheet.

Fast path (default): pyxlsb -> Reports/Valuable_Data_*.xlsx (~1 min)
Optional: --into-xlsb adds sheet "Valuable Data" inside the master .xlsb via Excel COM
           (close the .xlsb first; can take 5-15 min to open/save).
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import time
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

XLSB_NAME = "Copy of Copy of PSR_RAWDATA - new.xlsb"
DEST_SHEET = "Valuable Data"

FULL_SHEET_SECTIONS: tuple[tuple[str, str], ...] = (
    ("PSR&AVAIL", "PSR vs AVAIL (yearly trend — col C = manual AVAIL)"),
    ("BSC&RNC", "RNC 60-day chart (BSC&RNC)"),
    ("RNC", "RNC daily pivot"),
    ("BSC", "BSC daily pivot"),
    ("LAC BSC_RNC", "LAC performance"),
    ("Raw data for Avail Vs PSR", "Yearly network PSR source"),
)

# 1-based inclusive Excel ranges on 'Raw data - BSC_RNC' (skip full 289k-row sheet).
RAW_BSC_RNC_BLOCKS: tuple[tuple[str, str], ...] = (
    ("A1:N44", "Raw paste — RNC block"),
    ("A45:P90", "Raw paste — BSC block"),
    ("Q45:Z200", "Raw paste — LAC block"),
    ("T2:Y40", "Raw paste — Technology PSR (2G/3G/4G)"),
    ("AA2:AF40", "Raw paste — Service PSR"),
)

SHEET_RAW_BSC_RNC = "Raw data - BSC_RNC"
XL_PASTE_VALUES = -4163

_COL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _root() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _xlsb_path() -> str:
    return os.path.join(_root(), XLSB_NAME)


def _default_out() -> str:
    reports = os.path.join(_root(), "Reports")
    os.makedirs(reports, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return os.path.join(reports, f"Valuable_Data_{stamp}.xlsx")


def _backup_path(xlsb: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(xlsb)
    return f"{base}_backup_{stamp}{ext}"


def _col_to_index(col: str) -> int:
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _parse_range(address: str) -> tuple[int, int, int, int]:
    """Excel range -> 0-based inclusive row/col bounds."""
    a, b = address.split(":")
    m1, m2 = _COL_RE.match(a), _COL_RE.match(b)
    if not m1 or not m2:
        raise ValueError(f"Bad range: {address}")
    c1, r1 = _col_to_index(m1.group(1)), int(m1.group(2)) - 1
    c2, r2 = _col_to_index(m2.group(1)), int(m2.group(2)) - 1
    return r1, r2, c1, c2


def _row_to_list(row, width: int) -> list:
    out = [None] * width
    for cell in row:
        if cell.c < width:
            out[cell.c] = cell.v
    return out


def _read_sheet_all(xlsb: str, sheet_name: str) -> list[list]:
    from pyxlsb import open_workbook

    rows: list[list] = []
    max_col = 0
    with open_workbook(xlsb) as wb:
        with wb.get_sheet(sheet_name) as sh:
            for row in sh.rows():
                if not row:
                    continue
                max_col = max(max_col, max(c.c for c in row) + 1)
                rows.append(_row_to_list(row, max_col))

    # Trim trailing empty rows/cols.
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()
    if not rows:
        return []
    last_col = 0
    for r in rows:
        for i, v in enumerate(r):
            if v not in (None, ""):
                last_col = max(last_col, i + 1)
    return [r[:last_col] for r in rows]


def _read_sheet_range(xlsb: str, sheet_name: str, address: str) -> list[list]:
    from pyxlsb import open_workbook

    r1, r2, c1, c2 = _parse_range(address)
    width = c2 - c1 + 1
    want: dict[int, list] = {}
    with open_workbook(xlsb) as wb:
        with wb.get_sheet(sheet_name) as sh:
            for row in sh.rows():
                if not row:
                    continue
                r = row[0].r
                if r < r1:
                    continue
                if r > r2:
                    break
                line = [None] * width
                for cell in row:
                    if c1 <= cell.c <= c2:
                        line[cell.c - c1] = cell.v
                want[r] = line

    return [want[r] for r in range(r1, r2 + 1) if r in want]


def _write_sections_openpyxl(out: str, sections: list[tuple[str, list[list]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = DEST_SHEET
    row = 1
    ws.cell(row, 1, "PSR valuable data export").font = Font(bold=True)
    row = 3
    ws.cell(row, 1, f"Generated {datetime.now():%Y-%m-%d %H:%M}")
    row = 5

    for title, data in sections:
        ws.cell(row, 1, title).font = Font(bold=True)
        row += 1
        if not data:
            row += 1
            continue
        for i, line in enumerate(data):
            for j, val in enumerate(line, start=1):
                ws.cell(row + i, j, val)
        row += len(data) + 1

    wb.save(out)


def extract_pyxlsb(xlsb: str, out: str) -> list[tuple[str, list[list]]]:
    sections: list[tuple[str, list[list]]] = []

    for sheet_name, label in FULL_SHEET_SECTIONS:
        print(f"  Reading {sheet_name}...", flush=True)
        t0 = time.time()
        try:
            data = _read_sheet_all(xlsb, sheet_name)
        except Exception as exc:
            print(f"    WARNING: {exc}")
            sections.append((f"[MISSING] {label}", []))
            continue
        print(f"    {len(data):,} rows ({time.time() - t0:.0f}s)")
        sections.append((label, data))

    print(f"  Reading compact blocks from '{SHEET_RAW_BSC_RNC}'...", flush=True)
    for address, label in RAW_BSC_RNC_BLOCKS:
        t0 = time.time()
        data = _read_sheet_range(xlsb, SHEET_RAW_BSC_RNC, address)
        print(f"    {address}: {len(data):,} rows ({time.time() - t0:.0f}s)")
        sections.append((label, data))

    print(f"  Writing {out}...", flush=True)
    _write_sections_openpyxl(out, sections)
    size_mb = os.path.getsize(out) / 1024 / 1024
    print(f"  Saved ({size_mb:.2f} MB)")
    return sections


def inject_into_xlsb(xlsb: str, export_xlsx: str, *, backup: bool = True) -> None:
    if backup:
        bak = _backup_path(xlsb)
        print(f"Backup: {bak}")
        shutil.copy2(xlsb, bak)

    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    excel.AskToUpdateLinks = False

    try:
        print(f"Opening master .xlsb (close it in Excel/Cursor first)...", flush=True)
        t0 = time.time()
        master = excel.Workbooks.Open(os.path.abspath(xlsb), 0, False)
        print(f"  Opened in {time.time() - t0:.0f}s", flush=True)

        try:
            master.Worksheets(DEST_SHEET).Delete()
        except Exception:
            pass

        print(f"Importing sheet from {export_xlsx}...", flush=True)
        src = excel.Workbooks.Open(os.path.abspath(export_xlsx), 0, True)
        src.Worksheets(1).Copy(After=master.Worksheets(master.Worksheets.Count))
        master.Worksheets(master.Worksheets.Count).Name = DEST_SHEET
        master.Worksheets(DEST_SHEET).Tab.Color = 5296274
        master.Worksheets(DEST_SHEET).Activate()
        src.Close(False)

        print("Saving master workbook...", flush=True)
        t1 = time.time()
        master.Save()
        master.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()
        print(f"  Saved in {time.time() - t1:.0f}s")
        print(f"\nDone — sheet '{DEST_SHEET}' added to:\n  {xlsb}")
    except Exception:
        try:
            master.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()
        raise


def extract(
    xlsb: str,
    out: str,
    *,
    into_xlsb: bool = False,
    backup: bool = True,
) -> str:
    if not os.path.isfile(xlsb):
        raise FileNotFoundError(f"Not found: {xlsb}")

    print(f"Source: {xlsb} ({os.path.getsize(xlsb) / 1024 / 1024:.1f} MB)")
    print(f"Output: {out}\n")

    t0 = time.time()
    extract_pyxlsb(xlsb, out)
    print(f"\nExport finished in {time.time() - t0:.0f}s")

    if into_xlsb:
        print("\n--- Injecting into master .xlsb ---")
        inject_into_xlsb(xlsb, out, backup=backup)
    else:
        print(f"\nOpen: {out}")
        print(f"To also add sheet '{DEST_SHEET}' inside the .xlsb, run:")
        print(f'  extract_valuable.bat --into-xlsb')

    return out


def main() -> None:
    p = argparse.ArgumentParser(description="Extract valuable PSR data to consolidated sheet")
    p.add_argument("--xlsb", default=_xlsb_path())
    p.add_argument("--out", default=_default_out())
    p.add_argument(
        "--into-xlsb",
        action="store_true",
        help="Also add sheet 'Valuable Data' inside the master .xlsb (slow; close file first)",
    )
    p.add_argument("--no-backup", action="store_true")
    args = p.parse_args()
    extract(args.xlsb, args.out, into_xlsb=args.into_xlsb, backup=not args.no_backup)


if __name__ == "__main__":
    main()
