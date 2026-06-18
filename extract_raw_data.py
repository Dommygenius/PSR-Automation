"""
Extract ONLY raw data sheets from the master .xlsb into a light .xlsx for analysis.

Source sheets:
  - Raw data - BSC_RNC
  - Raw data for Avail Vs PSR
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

XLSB_NAME = "Copy of Copy of PSR_RAWDATA - new.xlsb"

SHEET_RAW_BSC_RNC = "Raw data - BSC_RNC"
SHEET_RAW_PSR_AVAIL = "Raw data for Avail Vs PSR"

EXPORT_SHEETS = (
    (SHEET_RAW_BSC_RNC, "Raw_BSC_RNC"),
    (SHEET_RAW_PSR_AVAIL, "Raw_PSR_AVAIL"),
)


def _root() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _xlsb_path() -> str:
    return os.path.join(_root(), XLSB_NAME)


def _default_out() -> str:
    reports = os.path.join(_root(), "Reports")
    os.makedirs(reports, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return os.path.join(reports, f"PSR_Raw_Data_{stamp}.xlsx")


def _write_rows_openpyxl(out: str, sheets: list[tuple[str, list[list]]]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        for i, row in enumerate(rows):
            if i and i % 50000 == 0:
                print(f"    writing row {i:,}...", flush=True)
            for j, val in enumerate(row, start=1):
                ws.cell(row=i + 1, column=j, value=val)
    print("  Saving .xlsx...", flush=True)
    wb.save(out)


def _com_range_to_rows(used) -> list[list]:
    """Convert Excel UsedRange COM object to list of lists."""
    val = used.Value
    if val is None:
        return []
    # Single cell
    if not isinstance(val, tuple):
        return [[val]]
    # Single row
    if not isinstance(val[0], tuple):
        return [list(val)]
    return [list(r) for r in val]


def _extract_excel_com(xlsb: str, out: str) -> None:
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    excel.AskToUpdateLinks = False

    try:
        print("  Opening .xlsb in Excel (often 2–5 min for large file)...", flush=True)
        t0 = time.time()
        src_wb = excel.Workbooks.Open(os.path.abspath(xlsb), 0, True)
        print(f"  Opened in {time.time() - t0:.0f}s", flush=True)

        out_sheets: list[tuple[str, list[list]]] = []
        for src_name, dst_name in EXPORT_SHEETS:
            print(f"  Reading '{src_name}'...", flush=True)
            try:
                ws = src_wb.Worksheets(src_name)
            except Exception:
                print(f"    WARNING: sheet '{src_name}' not found — skipped.")
                continue
            t1 = time.time()
            rows = _com_range_to_rows(ws.UsedRange)
            print(f"    -> {dst_name}: {len(rows):,} rows x {len(rows[0]) if rows else 0} cols ({time.time() - t1:.0f}s)")
            out_sheets.append((dst_name, rows))

        src_wb.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

        if not out_sheets:
            raise RuntimeError("No raw sheets found in workbook.")

        if os.path.isfile(out):
            os.remove(out)
        _write_rows_openpyxl(out, out_sheets)
    except Exception:
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()
        raise


def _extract_pyxlsb(xlsb: str, out: str) -> None:
    from pyxlsb import open_workbook

    out_sheets: list[tuple[str, list[list]]] = []
    for src_name, dst_name in EXPORT_SHEETS:
        print(f"  Reading '{src_name}' (pyxlsb — slow on big sheets)...", flush=True)
        rows: list[list] = []
        t0 = time.time()
        with open_workbook(xlsb) as wb:
            with wb.get_sheet(src_name) as sh:
                for i, row in enumerate(sh.rows(), start=1):
                    rows.append([c.v if c.v is not None else None for c in row])
                    if i % 50000 == 0:
                        print(f"    ... {i:,} rows read ({time.time() - t0:.0f}s)", flush=True)
        while rows and all(v is None or v == "" for v in rows[-1]):
            rows.pop()
        print(f"    -> {dst_name}: {len(rows):,} rows ({time.time() - t0:.0f}s)")
        out_sheets.append((dst_name, rows))

    if os.path.isfile(out):
        os.remove(out)
    _write_rows_openpyxl(out, out_sheets)


def extract(xlsb: str, out: str, *, method: str = "auto") -> None:
    if not os.path.isfile(xlsb):
        raise FileNotFoundError(f"Not found: {xlsb}")

    print(f"Source: {xlsb} ({os.path.getsize(xlsb) / 1024 / 1024:.1f} MB)")
    print(f"Output: {out}")
    print("Raw sheets only: Raw_BSC_RNC + Raw_PSR_AVAIL")
    print("")

    if method in ("auto", "com"):
        try:
            _extract_excel_com(xlsb, out)
            print(f"\nDone: {out} ({os.path.getsize(out) / 1024 / 1024:.1f} MB)")
            return
        except Exception as exc:
            if method == "com":
                raise
            print(f"\nExcel COM failed: {exc}")
            print("Falling back to pyxlsb...\n")

    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        raise RuntimeError("pip install pyxlsb   OR   use: python extract_raw_data.py --method com") from None

    _extract_pyxlsb(xlsb, out)
    print(f"\nDone: {out} ({os.path.getsize(out) / 1024 / 1024:.1f} MB)")


def main() -> None:
    p = argparse.ArgumentParser(description="Export raw RNC/BSC/PSR-AVAIL data from .xlsb")
    p.add_argument("--xlsb", default=_xlsb_path())
    p.add_argument("--out", default=_default_out())
    p.add_argument("--method", choices=("auto", "com", "pyxlsb"), default="auto")
    args = p.parse_args()
    extract(args.xlsb, args.out, method=args.method)


if __name__ == "__main__":
    main()
